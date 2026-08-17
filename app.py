import streamlit as st
import openpyxl
import pandas as pd
import re
import io
import os

st.set_page_config(
    page_title="중학교 자유학기 운영계획서 자동 검토 시스템",
    page_icon="📚",
    layout="wide"
)

# --- Helper Functions from Validator ---
def evaluate_formula_string(s):
    if not s:
        return 0.0
    cleaned = str(s).replace(',', '')
    cleaned = cleaned.replace('x', '*').replace('X', '*')
    units = ['원', '명', '회', '개', '박', '일', '명당', '회당', '식', '대', '명/회', '인']
    for u in units:
        cleaned = cleaned.replace(u, '')
    cleaned = re.sub(r'[^0-9\+\-\*\/\.\(\)]', '', cleaned)
    if not cleaned:
        return 0.0
    try:
        return float(eval(cleaned))
    except:
        return 0.0

def extract_nm_sum(cell_value):
    if not cell_value:
        return 0
    matches = re.findall(r'\((\d+)\)', str(cell_value))
    return sum(int(m) for m in matches)

def validate_free_semester_plan_bytes(file_bytes, filename):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    results = {"file": filename, "checks": [], "summary": ""}
    
    # 1. School Info
    if '1.학교운영 현황' not in wb.sheetnames:
        results["checks"].append({
            "category": "기본", "item": "시트 존재 여부", "result": "FAIL", "detail": "'1.학교운영 현황' 시트가 없습니다."
        })
        return results
        
    ws1 = wb['1.학교운영 현황']
    class_count = 0
    for row in ws1.iter_rows(min_row=1, max_row=15, min_col=1, max_col=10):
        for cell in row:
            if cell.value and '학급수' in str(cell.value):
                neighbor = ws1.cell(row=cell.row, column=cell.column+1).value
                if neighbor and str(neighbor).isdigit():
                    class_count = int(neighbor)
                break
                
    operation_hours = 0
    for row in ws1.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
        for cell in row:
            if cell.value and '자유학기 운영시수' in str(cell.value):
                val = ws1.cell(row=cell.row, column=cell.column+1).value
                if val and isinstance(val, (int, float)):
                    operation_hours = val
    if not operation_hours:
        for coord in ['F6', 'E6', 'F5', 'E5']:
            val = ws1[coord].value
            if isinstance(val, (int, float)) and val > 50:
                operation_hours = val
                break
                
    pass_1_1 = operation_hours >= 102
    results["checks"].append({
        "category": "1.학교운영 현황",
        "item": "자유학기 운영시수 >= 102",
        "result": "PASS" if pass_1_1 else "FAIL",
        "detail": f"운영시수: {operation_hours}시간"
    })
    
    d8_val = ws1['D8'].value or 0
    e8_val = ws1['E8'].value or ""
    sum_e8 = extract_nm_sum(e8_val)
    pass_1_2 = (int(d8_val) == sum_e8)
    results["checks"].append({
        "category": "1.학교운영 현황",
        "item": "주제선택 시수(D8)와 세부합(E8) 일치",
        "result": "PASS" if pass_1_2 else "FAIL",
        "detail": f"D8: {d8_val}, E8합: {sum_e8} ({e8_val})"
    })
    
    d9_val = ws1['D9'].value or 0
    e9_val = ws1['E9'].value or ""
    sum_e9 = extract_nm_sum(e9_val)
    pass_1_3 = (int(d9_val) == sum_e9)
    results["checks"].append({
        "category": "1.학교운영 현황",
        "item": "진로탐색 시수(D9)와 세부합(E9) 일치",
        "result": "PASS" if pass_1_3 else "FAIL",
        "detail": f"D9: {d9_val}, E9합: {sum_e9} ({e9_val})"
    })
    
    # 2. Activities
    if '2. 자유학기 활동' not in wb.sheetnames:
        results["checks"].append({
            "category": "기본", "item": "시트 존재 여부", "result": "FAIL", "detail": "'2. 자유학기 활동' 시트가 없습니다."
        })
        return results
        
    ws2 = wb['2. 자유학기 활동']
    subject_choice_hours_sum = 0
    career_search_hours_sum = 0
    current_cat = ""
    has_subject_personal = False
    has_career_personal = False
    
    for row in ws2.iter_rows(min_row=4, values_only=True):
        col1 = row[0]
        col2 = row[1]
        if col1 and not col2 and isinstance(col1, str) and "활동" in col1:
            current_cat = col1.strip()
            continue
        if isinstance(col1, (int, float)) or (isinstance(col1, str) and col1.isdigit()):
            program_name = row[1]
            instructor = row[4] if len(row) > 4 else ""
            sessions = row[5] if len(row) > 5 else 0
            total_hours = row[6] if len(row) > 6 else 0
            
            if instructor and "개인위탁" in str(instructor):
                if "주제선택" in current_cat:
                    has_subject_personal = True
                elif "진로" in current_cat:
                    has_career_personal = True
            
            if "주제선택" in current_cat:
                if total_hours and isinstance(total_hours, (int, float)):
                    subject_choice_hours_sum += total_hours
                if sessions and isinstance(sessions, (int, float)) and sessions < 2:
                    results["checks"].append({
                        "category": "2. 자유학기 활동",
                        "item": f"주제선택 프로그램 '{program_name}' 운영 회기 < 2",
                        "result": "FAIL",
                        "detail": f"회기: {sessions}"
                    })
            elif "진로" in current_cat:
                if total_hours and isinstance(total_hours, (int, float)):
                    career_search_hours_sum += total_hours

    target_subject_hours = int(d8_val) * class_count
    pass_2_1_1 = subject_choice_hours_sum >= target_subject_hours
    results["checks"].append({
        "category": "2. 자유학기 활동",
        "item": "주제선택 총 운영 시수 합계 검증",
        "result": "PASS" if pass_2_1_1 else "FAIL",
        "detail": f"합계: {subject_choice_hours_sum}, 기준(D8*학급수): {target_subject_hours} ({d8_val} * {class_count})"
    })
    
    target_career_hours = int(d9_val) * class_count
    pass_2_2_1 = career_search_hours_sum >= target_career_hours
    results["checks"].append({
        "category": "2. 자유학기 활동",
        "item": "진로탐색 총 운영 시수 합계 검증",
        "result": "PASS" if pass_2_2_1 else "FAIL",
        "detail": f"합계: {career_search_hours_sum}, 기준(D9*학급수): {target_career_hours} ({d9_val} * {class_count})"
    })
    
    # 3. Budget
    if '3. 예산 계획서' in wb.sheetnames:
        ws3 = wb['3. 예산 계획서']
        total_budget = ws3['E3'].value
        if not isinstance(total_budget, (int, float)):
            total_budget = 0
            for row in ws3.iter_rows(min_row=1, max_row=5, values_only=True):
                for idx, cell_val in enumerate(row):
                    if cell_val and '자유학기 총 예산' in str(cell_val):
                        for next_val in row[idx+1:]:
                            if isinstance(next_val, (int, float)) and next_val > 1000:
                                total_budget = next_val
                                break
                    if total_budget: break
                    
        budget_items_valid = True
        for row in ws3.iter_rows(min_row=5, values_only=True):
            content = row[1]
            formula = row[2]
            budget_val = row[3]
            if content and budget_val and isinstance(budget_val, (int, float)):
                calc_val = evaluate_formula_string(formula)
                if calc_val > 0 and abs(calc_val - budget_val) > 10:
                    results["checks"].append({
                        "category": "3. 예산 계획서",
                        "item": f"산출근거 불일치 ({content})",
                        "result": "FAIL",
                        "detail": f"수식: {formula} (계산: {calc_val}), 소요예산: {budget_val}"
                    })
                    budget_items_valid = False
        if budget_items_valid:
            results["checks"].append({
                "category": "3. 예산 계획서",
                "item": "산출근거 및 소요예산 일치 여부",
                "result": "PASS",
                "detail": "모든 예산 항목의 산출근거가 일치합니다."
            })
            
        biz_fee = 0
        for row in ws3.iter_rows(min_row=5, values_only=True):
            if row[0] and '업무추진비' in str(row[0]):
                if isinstance(row[3], (int, float)):
                    biz_fee = row[3]
        limit_3_2 = total_budget * 0.03
        pass_3_2 = biz_fee < limit_3_2
        results["checks"].append({
            "category": "3. 예산 계획서",
            "item": "업무추진비 3% 미만 여부",
            "result": "PASS" if pass_3_2 else "FAIL",
            "detail": f"업무추진비: {biz_fee:,}원, 3% 한도: {limit_3_2:,.0f}원 (총예산: {total_budget:,}원)"
        })
        
        personal_entrustment_budget = 0
        has_personal_item_in_budget = False
        for row in ws3.iter_rows(min_row=5, values_only=True):
            full_row_str = " ".join([str(c) for c in row if c])
            if '개인위탁' in full_row_str:
                has_personal_item_in_budget = True
                if isinstance(row[3], (int, float)):
                    personal_entrustment_budget += row[3]
                    
        any_personal_in_activity = has_subject_personal or has_career_personal
        if any_personal_in_activity:
            pass_3_3_1 = has_personal_item_in_budget
            results["checks"].append({
                "category": "3. 예산 계획서",
                "item": "활동 개인위탁 시 예산서 반영 여부",
                "result": "PASS" if pass_3_3_1 else "FAIL",
                "detail": f"활동 내 개인위탁 존재 여부 반영: {has_personal_item_in_budget}"
            })
            limit_ratio = 0.50 if has_career_personal else 0.40
            limit_amount = total_budget * limit_ratio
            pass_3_3_2 = personal_entrustment_budget <= limit_amount
            results["checks"].append({
                "category": "3. 예산 계획서",
                "item": f"개인위탁 비용 한도 준수 ({int(limit_ratio*100)}% 이내)",
                "result": "PASS" if pass_3_3_2 else "FAIL",
                "detail": f"개인위탁 비용: {personal_entrustment_budget:,}원, 한도: {limit_amount:,.0f}원"
            })
        else:
            results["checks"].append({
                "category": "3. 예산 계획서",
                "item": "개인위탁 관련 예산 검토",
                "result": "PASS",
                "detail": "자유학기 활동에 개인위탁 교사가 없습니다."
            })

    # Summary
    pass_count = sum(1 for c in results["checks"] if c["result"] == "PASS")
    fail_count = sum(1 for c in results["checks"] if c["result"] == "FAIL")
    results["pass_count"] = pass_count
    results["fail_count"] = fail_count
    
    if fail_count == 0:
        results["summary"] = "모든 점검 항목을 완벽하게 통과하였습니다. 특이사항이 없습니다."
    else:
        results["summary"] = f"총 {fail_count건의 보완 필요 사항(불일치/기준 미달)이 발견되었습니다. 상세 내용을 확인 후 수정이 필요합니다."
        
    return results

# --- Streamlit UI ---
st.title("📚 중학교 자유학기 운영계획서 자동 검토 웹앱")
st.markdown("교육청 표준 서식으로 작성된 자유학기 운영계획서 엑셀 파일을 일괄 업로드하여 자동으로 검토하고 결과를 확인하는 시스템입니다.")

# Sidebar for controls
st.sidebar.header("⚙️ 제어판")

# Batch reset button using session state
if 'file_uploader_key' not in st.sidebar:
    st.sidebar.file_uploader_key = 0

def reset_uploader():
    st.sidebar.file_uploader_key += 1

st.sidebar.button("🔄 업로드 파일 일괄 초기화", on_click=reset_uploader)

uploaded_files = st.file_uploader(
    "자유학기 운영계획서 엑셀 파일들을 업로드하세요 (복수 선택 가능)",
    type=["xlsx"],
    accept_multiple_files=True,
    key=f"file_uploader_{st.sidebar.file_uploader_key}"
)

if uploaded_files:
    st.success(f"총 {len(uploaded_files)}개의 파일이 업로드되었습니다.")
    
    # Process all files
    all_results = []
    tabs = st.tabs([f"{f.name.split('_')[0]}중" for f in uploaded_files])
    
    # Excel export buffer
    output_excel = io.BytesIO()
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for idx, uploaded_file in enumerate(uploaded_files):
            file_bytes = uploaded_file.read()
            res = validate_free_semester_plan_bytes(file_bytes, uploaded_file.name)
            all_results.append(res)
            
            with tabs[idx]:
                st.subheader(f"🏫 {uploaded_file.name} 검토 결과")
                
                # Total Summary Box
                if res['fail_count'] == 0:
                    st.success(f"**총평**: {res['summary']} (PASS: {res['pass_count']} / FAIL: {res['fail_count']})")
                else:
                    st.error(f"**총평**: {res['summary']} (PASS: {res['pass_count']} / FAIL: {res['fail_count']})")
                
                # Discrepancies summary if any
                fails = [c for c in res['checks'] if c['result'] == 'FAIL']
                if fails:
                    st.markdown("### ⚠️ 주요 불일치 및 보완 필요 사항")
                    for f_item in fails:
                        st.warning(f"**[{f_item['category']}] {f_item['item']}**\n- 상세: {f_item['detail']}")
                
                # Full Check Table
                st.markdown("### 📋 전체 점검 항목 상세 내역")
                df_checks = pd.DataFrame(res['checks'])
                # Map PASS/FAIL to emoji
                df_checks['상태'] = df_checks['result'].apply(lambda x: '✅ PASS' if x == 'PASS' else '❌ FAIL')
                df_display = df_checks[['상태', 'category', 'item', 'detail']]
                df_display.columns = ['상태', '검토 영역', '점검 항목', '상세 내용']
                st.dataframe(df_display, use_container_width=True)
                
            # Write sheet for excel export
            df_sheet = pd.DataFrame(res['checks'])
            sheet_name = uploaded_file.name[:25].replace('[', '').replace(']', '').replace('*', '')
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
            
    output_excel.seek(0)
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("📥 결과 다운로드")
    st.sidebar.download_button(
        label="📊 전체 학교 검토 결과 엑셀 다운로드",
        data=output_excel,
        file_name="자유학기_계획서_일괄검토결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("좌측 또는 위 업로드 박스에 자유학기 운영계획서 엑셀 파일을 업로드해 주세요.")
